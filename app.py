from flask import Flask, request, jsonify, render_template, redirect, session as flask_session, render_template_string, send_file
from flask_cors import CORS
import sqlite3
import datetime
import secrets
import io
import os
from werkzeug.security import generate_password_hash, check_password_hash

# Deteksi otomatis letak template HTML secara absolut terhadap file app.py
base_dir = os.path.dirname(os.path.abspath(__file__))
templates_path = os.path.join(base_dir, 'templates')
template_dir = templates_path if os.path.isdir(templates_path) else base_dir

app = Flask(__name__, template_folder=template_dir)

app.secret_key = secrets.token_hex(32)
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024 # 10 MB limit
DATABASE = 'school_db.sqlite'

# Helper untuk mendapatkan koneksi database dengan Foreign Key aktif
def get_db_connection():
    conn = sqlite3.connect(DATABASE)
    conn.execute("PRAGMA foreign_keys = ON;")
    return conn

def format_indo_date():
    months = ["Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
    now = datetime.date.today()
    return f"{now.day} {months[now.month - 1]} {now.year}"

# Merakit nama guru dengan gelar secara profesional
def get_full_name(nama, depan, belakang):
    full = ""
    if depan and depan.strip():
        full += depan.strip() + " "
    full += (nama or "").strip()
    if belakang and belakang.strip():
        full += ", " + belakang.strip()
    return full.strip()

# ==========================================
# 🗄️ DATABASE INITIALIZATION & MIGRATIONS
# ==========================================
def init_db():
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # 1. Master Mapel
    cursor.execute('''CREATE TABLE IF NOT EXISTS mapels (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nama_mapel TEXT UNIQUE NOT NULL,
        kode_mapel TEXT DEFAULT ''
    )''')
    
    # 2. Master Kelas
    cursor.execute('''CREATE TABLE IF NOT EXISTS classes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nama_kelas TEXT UNIQUE NOT NULL
    )''')
    
    # 3. Master Users
    cursor.execute('''CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        role TEXT DEFAULT 'guru',
        tipe_guru TEXT DEFAULT 'mapel',
        nama_guru TEXT,
        nip TEXT DEFAULT '1234567890',
        gelar_depan TEXT DEFAULT '',
        gelar_belakang TEXT DEFAULT ''
    )''')

    # Bikin akun admin default jika belum ada
    cursor.execute("SELECT id FROM users WHERE username='admin'")
    if not cursor.fetchone():
        cursor.execute("INSERT INTO users (username, password_hash, role, nama_guru, nip, gelar_depan, gelar_belakang) VALUES (?, ?, ?, ?, ?, ?, ?)",
                       ('admin', generate_password_hash('admin123'), 'admin', 'Administrator', '1234567890', '', ''))

    # 4. Tabel Pivot (Jadwal Mengajar)
    cursor.execute('''CREATE TABLE IF NOT EXISTS jadwal_mengajar (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        class_id INTEGER,
        mapel_id INTEGER,
        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
        FOREIGN KEY (class_id) REFERENCES classes(id) ON DELETE CASCADE,
        FOREIGN KEY (mapel_id) REFERENCES mapels(id) ON DELETE CASCADE
    )''')
    
    # 5. Master Siswa
    cursor.execute('''CREATE TABLE IF NOT EXISTS students (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        class_id INTEGER,
        nama_siswa TEXT NOT NULL,
        nisn VARCHAR(20) UNIQUE NOT NULL,
        gender TEXT DEFAULT 'L',
        FOREIGN KEY (class_id) REFERENCES classes(id) ON DELETE CASCADE
    )''')
    
    # 6. Nilai Akhir
    cursor.execute('''CREATE TABLE IF NOT EXISTS nilai_akhir (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id INTEGER,
        mapel_id INTEGER,
        nilai_asli REAL,
        nilai_ai REAL,
        nh1 REAL,
        nh2 REAL,
        nh3 REAL,
        tanggal_input DATETIME DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(student_id, mapel_id),
        FOREIGN KEY (student_id) REFERENCES students(id) ON DELETE CASCADE,
        FOREIGN KEY (mapel_id) REFERENCES mapels(id) ON DELETE CASCADE
    )''')

    # Skema Migrasi aman untuk DB yang sudah ada
    try:
        cursor.execute("PRAGMA table_info(students)")
        cols = [c[1] for c in cursor.fetchall()]
        if 'gender' not in cols:
            cursor.execute("ALTER TABLE students ADD COLUMN gender TEXT DEFAULT 'L'")
    except Exception as e:
        print(f"Migration students info skipped: {e}")

    try:
        cursor.execute("PRAGMA table_info(users)")
        cols = [c[1] for c in cursor.fetchall()]
        if 'nip' not in cols:
            cursor.execute("ALTER TABLE users ADD COLUMN nip TEXT DEFAULT '1234567890'")
        if 'gelar_depan' not in cols:
            cursor.execute("ALTER TABLE users ADD COLUMN gelar_depan TEXT DEFAULT ''")
        if 'gelar_belakang' not in cols:
            cursor.execute("ALTER TABLE users ADD COLUMN gelar_belakang TEXT DEFAULT ''")
    except Exception as e:
        print(f"Migration users info skipped: {e}")

    try:
        cursor.execute("PRAGMA table_info(mapels)")
        cols = [c[1] for c in cursor.fetchall()]
        if 'kode_mapel' not in cols:
            cursor.execute("ALTER TABLE mapels ADD COLUMN kode_mapel TEXT DEFAULT ''")
    except Exception as e:
        print(f"Migration mapels failed: {e}")

    try:
        cursor.execute("PRAGMA table_info(nilai_akhir)")
        cols = [c[1] for c in cursor.fetchall()]
        if 'nh1' not in cols:
            cursor.execute("ALTER TABLE nilai_akhir ADD COLUMN nh1 REAL")
        if 'nh2' not in cols:
            cursor.execute("ALTER TABLE nilai_akhir ADD COLUMN nh2 REAL")
        if 'nh3' not in cols:
            cursor.execute("ALTER TABLE nilai_akhir ADD COLUMN nh3 REAL")
    except Exception as e:
        print(f"Migration nilai_akhir failed: {e}")

    conn.commit()
    conn.close()

try:
    init_db()
except Exception as e:
    print(f"⚠️ Error Database: {e}")


# ==========================================
# 🌐 ROUTING HALAMAN HTML (RENDER TEMPLATE)
# ==========================================
@app.route('/')
def index():
    if 'user_id' not in flask_session:
        return redirect('/login')
    if flask_session.get('role') == 'admin':
        return redirect('/admin')
    else:
        return redirect('/guru')

@app.route('/login')
def login_page():
    if 'user_id' in flask_session:
        return redirect('/')
    return render_template('login.html')

@app.route('/admin')
def admin_page():
    if flask_session.get('role') != 'admin':
        return redirect('/')
    return render_template('admin.html', nama=flask_session.get('nama_guru'))

@app.route('/guru')
def guru_page():
    if flask_session.get('role') != 'guru':
        return redirect('/')
    return render_template('guru.html', nama=flask_session.get('nama_guru'), tipe=flask_session.get('tipe_guru'))


# ==========================================
# 🔐 API AUTHENTICATION
# ==========================================
@app.route('/api/login', methods=['POST'])
def api_login():
    data = request.get_json() or {}
    username = data.get('username', '').strip().lower()
    password = data.get('password', '')
    
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT id, username, password_hash, role, tipe_guru, nama_guru, nip, gelar_depan, gelar_belakang FROM users WHERE username = ? LIMIT 1', (username,))
    row = cursor.fetchone()
    conn.close()
    
    if not row or not check_password_hash(row[2], password):
        return jsonify({'success': False, 'message': 'Username atau password salah!'}), 401
    
    flask_session['user_id'] = row[0]
    flask_session['username'] = row[1]
    flask_session['role'] = row[3]
    flask_session['tipe_guru'] = row[4]
    flask_session['nama_guru'] = get_full_name(row[5], row[7], row[8])
    flask_session['nip'] = row[6]
    
    return jsonify({'success': True, 'redirect': '/'})

@app.route('/api/logout', methods=['POST', 'GET'])
def api_logout():
    flask_session.clear()
    return redirect('/login')

# ==========================================
# 🛠️ API ADMIN (CRUD SISWA, MAPEL, KELAS, & GURU)
# ==========================================

# 1. API CRUD SISWA
@app.route('/api/admin/students', methods=['GET', 'POST'])
def api_students():
    if flask_session.get('role') != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    conn = get_db_connection()
    cursor = conn.cursor()
    if request.method == 'GET':
        class_id = request.args.get('class_id')
        if class_id:
            cursor.execute('''
                SELECT s.id, s.class_id, c.nama_kelas, s.nama_siswa, s.nisn, s.gender 
                FROM students s
                JOIN classes c ON s.class_id = c.id
                WHERE s.class_id = ?
                ORDER BY s.nama_siswa ASC
            ''', (class_id,))
        else:
            cursor.execute('''
                SELECT s.id, s.class_id, c.nama_kelas, s.nama_siswa, s.nisn, s.gender 
                FROM students s
                JOIN classes c ON s.class_id = c.id
                ORDER BY c.nama_kelas ASC, s.nama_siswa ASC
            ''')
        students = [{'id': r[0], 'class_id': r[1], 'nama_kelas': r[2], 'nama_siswa': r[3], 'nisn': r[4], 'gender': r[5]} for r in cursor.fetchall()]
        conn.close()
        return jsonify(students)
        
    elif request.method == 'POST':
        data = request.get_json() or {}
        class_id = data.get('class_id')
        nama_siswa = data.get('nama_siswa', '').strip()
        nisn = data.get('nisn', '').strip()
        gender = data.get('gender', 'L').strip().upper()
        
        if not all([class_id, nama_siswa, nisn]):
            return jsonify({'success': False, 'message': 'Lengkapi isian nama, kelas, dan NISN'}), 400
        try:
            cursor.execute('INSERT INTO students (class_id, nama_siswa, nisn, gender) VALUES (?, ?, ?, ?)',
                           (class_id, nama_siswa, nisn, gender))
            conn.commit()
            conn.close()
            return jsonify({'success': True, 'message': 'Siswa berhasil ditambahkan!'})
        except sqlite3.IntegrityError:
            return jsonify({'success': False, 'message': 'NISN tersebut sudah digunakan siswa lain!'}), 400

@app.route('/api/admin/students/<int:student_id>', methods=['PUT', 'DELETE'])
def api_student_detail(student_id):
    if flask_session.get('role') != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    conn = get_db_connection()
    cursor = conn.cursor()
    if request.method == 'PUT':
        data = request.get_json() or {}
        class_id = data.get('class_id')
        nama_siswa = data.get('nama_siswa', '').strip()
        nisn = data.get('nisn', '').strip()
        gender = data.get('gender', 'L').strip().upper()
        
        if not all([class_id, nama_siswa, nisn]):
            return jsonify({'success': False, 'message': 'Lengkapi isian nama, kelas, dan NISN'}), 400
        try:
            cursor.execute('UPDATE students SET class_id=?, nama_siswa=?, nisn=?, gender=? WHERE id=?',
                           (class_id, nama_siswa, nisn, gender, student_id))
            conn.commit()
            conn.close()
            return jsonify({'success': True, 'message': 'Siswa berhasil diperbarui!'})
        except sqlite3.IntegrityError:
            return jsonify({'success': False, 'message': 'NISN tersebut sudah digunakan siswa lain!'}), 400
    elif request.method == 'DELETE':
        cursor.execute('DELETE FROM students WHERE id = ?', (student_id,))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Siswa berhasil dihapus'})

@app.route('/api/admin/students/clear', methods=['DELETE'])
def api_students_clear():
    if flask_session.get('role') != 'admin':
         return jsonify({'error': 'Unauthorized'}), 403
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM students')
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'message': 'Seluruh basis data siswa berhasil dikosongkan!'})

@app.route('/api/admin/students/import', methods=['POST'])
def api_students_import():
    if flask_session.get('role') != 'admin':
         return jsonify({'error': 'Unauthorized'}), 403
    if 'file' not in request.files:
         return jsonify({'success': False, 'message': 'File berkas tidak terunggah'}), 400
         
    file = request.files['file']
    class_id = request.form.get('class_id')
    if not class_id:
         return jsonify({'success': False, 'message': 'Pilih kelas penampung terlebih dahulu!'}), 400
         
    try:
        import pandas as pd
        df = pd.read_excel(file)
        
        # Mapping kolom secara mandiri
        cols_map = {}
        for col in df.columns:
            c_low = str(col).lower().strip()
            if any(x in c_low for x in ['nama', 'name', 'siswa', 'student']):
                cols_map['nama'] = col
            elif 'nisn' in c_low or 'nis' in c_low:
                cols_map['nisn'] = col
            elif any(x in c_low for x in ['gender', 'kelamin', 'lp', 'l/p', 'sex']):
                cols_map['gender'] = col
                
        if 'nama' not in cols_map or 'nisn' not in cols_map:
            return jsonify({'success': False, 'message': 'Berkas Excel wajib mengandung kolom Nama dan NISN!'}), 400
            
        conn = get_db_connection()
        cursor = conn.cursor()
        
        count = 0
        for _, row in df.iterrows():
            nama = str(row[cols_map['nama']]).strip()
            nisn = str(row[cols_map['nisn']]).strip()
            
            if not nama or not nisn or pd.isna(row[cols_map['nama']]) or pd.isna(row[cols_map['nisn']]):
                continue
                
            if nisn.endswith('.0'):
                nisn = nisn[:-2]
                
            gender = 'L'
            if 'gender' in cols_map:
                g_val = str(row[cols_map['gender']]).strip().upper()
                if any(x in g_val for x in ['P', 'W', 'FEMALE', 'PR']):
                    gender = 'P'
            
            # Melakukan INSERT atau REPLACE jika NISN bentrok
            cursor.execute('''
                INSERT INTO students (class_id, nama_siswa, nisn, gender) 
                VALUES (?, ?, ?, ?)
                ON CONFLICT(nisn) DO UPDATE SET 
                    nama_siswa=excluded.nama_siswa, 
                    class_id=excluded.class_id, 
                    gender=excluded.gender
            ''', (class_id, nama, nisn, gender))
            count += 1
            
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': f'Berhasil mengimpor/memperbarui {count} siswa!'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Internal Server Error: {str(e)}'}), 500

# 2. MASTER MAPEL
@app.route('/api/admin/mapels', methods=['GET', 'POST'])
def api_mapels():
    if flask_session.get('role') != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    conn = get_db_connection()
    cursor = conn.cursor()
    if request.method == 'GET':
        cursor.execute('SELECT id, nama_mapel, kode_mapel FROM mapels ORDER BY nama_mapel ASC')
        mapels = [{'id': row[0], 'nama_mapel': row[1], 'kode_mapel': row[2] if row[2] else ''} for row in cursor.fetchall()]
        conn.close()
        return jsonify(mapels)
    elif request.method == 'POST':
        data = request.get_json()
        nama_mapel = data.get('nama_mapel', '').strip()
        kode_mapel = data.get('kode_mapel', '').strip().upper()
        if not nama_mapel or not kode_mapel:
            return jsonify({'success': False, 'message': 'Nama dan Kode mapel tidak boleh kosong'}), 400
        try:
            cursor.execute('INSERT INTO mapels (nama_mapel, kode_mapel) VALUES (?, ?)', (nama_mapel, kode_mapel))
            conn.commit()
            conn.close()
            return jsonify({'success': True, 'message': 'Mapel berhasil ditambahkan!'})
        except sqlite3.IntegrityError:
            return jsonify({'success': False, 'message': 'Nama atau Kode mapel tersebut sudah ada!'}), 400

@app.route('/api/admin/mapels/<int:mapel_id>', methods=['DELETE'])
def delete_mapel(mapel_id):
    if flask_session.get('role') != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM mapels WHERE id = ?', (mapel_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'message': 'Mapel berhasil dihapus'})

# 3. MASTER KELAS
@app.route('/api/admin/classes', methods=['GET', 'POST'])
def api_classes():
    if flask_session.get('role') != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    conn = get_db_connection()
    cursor = conn.cursor()
    if request.method == 'GET':
        cursor.execute('SELECT id, nama_kelas FROM classes ORDER BY nama_kelas ASC')
        classes = [{'id': row[0], 'nama_kelas': row[1]} for row in cursor.fetchall()]
        conn.close()
        return jsonify(classes)
    elif request.method == 'POST':
        data = request.get_json()
        nama_kelas = data.get('nama_kelas', '').strip()
        if not nama_kelas:
            return jsonify({'success': False, 'message': 'Nama kelas tidak boleh kosong'}), 400
        try:
            cursor.execute('INSERT INTO classes (nama_kelas) VALUES (?)', (nama_kelas,))
            conn.commit()
            conn.close()
            return jsonify({'success': True, 'message': 'Kelas berhasil ditambahkan!'})
        except sqlite3.IntegrityError:
            return jsonify({'success': False, 'message': 'Kelas tersebut sudah ada!'}), 400

@app.route('/api/admin/classes/<int:class_id>', methods=['DELETE'])
def delete_class(class_id):
    if flask_session.get('role') != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM classes WHERE id = ?', (class_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'message': 'Kelas berhasil dihapus'})

# 4. MASTER GURU
@app.route('/api/admin/gurus', methods=['GET', 'POST'])
def api_gurus():
    if flask_session.get('role') != 'admin': return jsonify({'error': 'Unauthorized'}), 403
    conn = get_db_connection()
    cursor = conn.cursor()
    if request.method == 'GET':
        cursor.execute("SELECT id, username, nama_guru, tipe_guru, nip, "
                       "COALESCE(gelar_depan, '') as gd, COALESCE(gelar_belakang, '') as gb "
                       "FROM users WHERE role='guru' ORDER BY id DESC")
        gurus = []
        for row in cursor.fetchall():
            display_name = get_full_name(row[2], row[5], row[6])
            gurus.append({
                'id': row[0], 'username': row[1], 'nama_guru': row[2], 
                'nama_guru_lengkap': display_name, 'tipe_guru': row[3], 'nip': row[4],
                'gelar_depan': row[5], 'gelar_belakang': row[6]
            })
        conn.close()
        return jsonify(gurus)
    elif request.method == 'POST':
        data = request.get_json() or {}
        username = data.get('username', '').strip().lower()
        password = data.get('password', '')
        nama_guru = data.get('nama_guru', '').strip()
        tipe_guru = data.get('tipe_guru', 'mapel')
        nip = data.get('nip', '').strip() or '1234567890'
        gelar_depan = data.get('gelar_depan', '').strip()
        gelar_belakang = data.get('gelar_belakang', '').strip()
        
        if not username or not password:
            return jsonify({'success': False, 'message': 'Username dan password wajib diisi'}), 400
        try:
            password_hash = generate_password_hash(password)
            cursor.execute('''INSERT INTO users (username, password_hash, role, tipe_guru, nama_guru, nip, gelar_depan, gelar_belakang) 
                              VALUES (?, ?, 'guru', ?, ?, ?, ?, ?)''', (username, password_hash, tipe_guru, nama_guru, nip, gelar_depan, gelar_belakang))
            conn.commit()
            conn.close()
            return jsonify({'success': True, 'message': 'Guru berhasil ditambahkan!'})
        except sqlite3.IntegrityError as e:
            print(f"Error integrity insert guru: {e}")
            return jsonify({'success': False, 'message': 'Username sudah digunakan!'}), 400

@app.route('/api/admin/gurus/<int:guru_id>', methods=['DELETE'])
def delete_guru(guru_id):
    if flask_session.get('role') != 'admin': return jsonify({'error': 'Unauthorized'}), 403
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM users WHERE id = ? AND role = 'guru'", (guru_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'message': 'Guru berhasil dihapus'})

# 5. PLOTTING JADWAL
@app.route('/api/admin/jadwal', methods=['GET', 'POST'])
def api_jadwal():
    if flask_session.get('role') != 'admin': return jsonify({'error': 'Unauthorized'}), 403
    conn = get_db_connection()
    cursor = conn.cursor()
    if request.method == 'GET':
        cursor.execute('''
            SELECT j.id, u.nama_guru, c.nama_kelas, m.nama_mapel, u.tipe_guru, u.gelar_depan, u.gelar_belakang
            FROM jadwal_mengajar j
            JOIN users u ON j.user_id = u.id
            JOIN classes c ON j.class_id = c.id
            JOIN mapels m ON j.mapel_id = m.id
            ORDER BY u.nama_guru ASC, c.nama_kelas ASC
        ''')
        jadwal = []
        for r in cursor.fetchall():
            display_name = get_full_name(r[1], r[5], r[6])
            jadwal.append({'id': r[0], 'nama_guru': display_name, 'nama_kelas': r[2], 'nama_mapel': r[3], 'tipe_guru': r[4]})
        conn.close()
        return jsonify(jadwal)
    elif request.method == 'POST':
        data = request.get_json()
        user_id = data.get('user_id')
        class_id = data.get('class_id')
        mapel_id = data.get('mapel_id')
        if not all([user_id, class_id, mapel_id]):
            return jsonify({'success': False, 'message': 'Guru, Kelas, dan Mapel harus dipilih semua!'}), 400
        cursor.execute("SELECT id FROM jadwal_mengajar WHERE user_id=? AND class_id=? AND mapel_id=?", (user_id, class_id, mapel_id))
        if cursor.fetchone():
            conn.close()
            return jsonify({'success': False, 'message': 'Plotting ini sudah ada!'}), 400
        cursor.execute("INSERT INTO jadwal_mengajar (user_id, class_id, mapel_id) VALUES (?, ?, ?)", (user_id, class_id, mapel_id))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Plotting berhasil!'})

@app.route('/api/admin/jadwal/<int:jadwal_id>', methods=['DELETE'])
def delete_jadwal(jadwal_id):
    if flask_session.get('role') != 'admin': return jsonify({'error': 'Unauthorized'}), 403
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM jadwal_mengajar WHERE id = ?", (jadwal_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# 6. MONITORING
@app.route('/api/admin/monitoring', methods=['GET'])
def api_monitoring():
    if flask_session.get('role') != 'admin': 
        return jsonify({'error': 'Unauthorized'}), 403
    class_id = request.args.get('class_id')
    mapel_id = request.args.get('mapel_id')
    if not class_id or not mapel_id: 
        return jsonify([])
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT s.nama_siswa, s.nisn, n.nilai_asli, n.nilai_ai, s.gender,
               RANK() OVER (ORDER BY n.nilai_ai DESC) as ranking
        FROM students s
        JOIN nilai_akhir n ON s.id = n.student_id
        WHERE s.class_id = ? AND n.mapel_id = ?
        ORDER BY ranking ASC
    ''', (class_id, mapel_id))
    data = [{'nama_siswa': r[0], 'nisn': r[1], 'nilai_asli': r[2], 'nilai_ai': r[3], 'gender': r[4], 'ranking': r[5]} for r in cursor.fetchall()]
    conn.close()
    return jsonify(data)


# ==========================================
# 🎓 API GURU (JADWAL, TEMPLATE, UPLOAD, REKAP, & MANUAL INPUT)
# ==========================================
@app.route('/api/guru/assignments', methods=['GET'])
def get_assignments():
    if flask_session.get('role') != 'guru': 
        return jsonify({'error': 'Unauthorized'}), 403
        
    user_id = flask_session.get('user_id')
    tipe_guru = flask_session.get('tipe_guru')  # Mengambil tipe 'kelas' atau 'mapel'
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    if tipe_guru == 'kelas':
        # Untuk GURU KELAS:
        # 1. Ambil semua kelas yang ditugaskan kepadanya di tabel jadwal_mengajar
        cursor.execute('''
            SELECT DISTINCT c.id, c.nama_kelas
            FROM jadwal_mengajar j
            JOIN classes c ON j.class_id = c.id
            WHERE j.user_id = ?
        ''', (user_id,))
        assigned_classes = cursor.fetchall()
        
        # 2. Ambil seluruh mata pelajaran yang terdaftar di database sekolah
        cursor.execute('SELECT id, nama_mapel FROM mapels ORDER BY nama_mapel ASC')
        all_mapels = cursor.fetchall()
        
        # 3. Gabungkan kelas dengan seluruh mapel secara otomatis (Cross Join lokal)
        assignments = []
        for class_id, class_name in assigned_classes:
            for mapel_id, mapel_name in all_mapels:
                assignments.append({
                    'class_id': class_id,
                    'nama_kelas': class_name,
                    'mapel_id': mapel_id,
                    'nama_mapel': mapel_name
                })
    else:
        # Untuk GURU MAPEL:
        # Ambil hanya kelas dan mata pelajaran spesifik yang ditugaskan oleh Admin
        cursor.execute('''
            SELECT c.id as class_id, c.nama_kelas, m.id as mapel_id, m.nama_mapel
            FROM jadwal_mengajar j
            JOIN classes c ON j.class_id = c.id
            JOIN mapels m ON j.mapel_id = m.id
            WHERE j.user_id = ?
            ORDER BY c.nama_kelas ASC, m.nama_mapel ASC
        ''', (user_id,))
        assignments = [{
            'class_id': row[0], 
            'nama_kelas': row[1], 
            'mapel_id': row[2], 
            'nama_mapel': row[3]
        } for row in cursor.fetchall()]
        
    conn.close()
    return jsonify(assignments)

# Download template Excel yang sudah terisi data siswa kelas tersebut
@app.route('/api/guru/template', methods=['GET'])
def download_template():
    if flask_session.get('role') != 'guru':
        return jsonify({'error': 'Unauthorized'}), 403
    class_id = request.args.get('class_id')
    if not class_id:
        return "class_id wajib diisi", 400
        
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT nama_kelas FROM classes WHERE id=?", (class_id,))
    class_row = cursor.fetchone()
    nama_kelas = class_row[0] if class_row else f"Kelas_{class_id}"
    
    cursor.execute("SELECT nama_siswa, nisn, gender FROM students WHERE class_id = ? ORDER BY nama_siswa ASC", (class_id,))
    rows = cursor.fetchall()
    conn.close()
    
    import pandas as pd
    data = []
    if rows:
        for r in rows:
            data.append({
                'NISN': r[1],
                'Nama Siswa': r[0],
                'L/P': r[2],
                'NH-1': "",
                'NH-2': "",
                'NH-3': ""
            })
    else:
        # Template dummy jika belum ada data siswa di kelas ini
        for i in range(1, 6):
            data.append({
                'NISN': f"MOCK_NISN_{i}",
                'Nama Siswa': f"Siswa Baru {i}",
                'L/P': "L",
                'NH-1': "",
                'NH-2': "",
                'NH-3': ""
            })
            
    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Template Nilai')
    output.seek(0)
    
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"Template_Nilai_{nama_kelas}.xlsx"
    )

# Input manual nilai dari layar
@app.route('/api/guru/save_manual', methods=['POST'])
def save_manual_grades():
    if flask_session.get('role') != 'guru':
        return jsonify({'error': 'Unauthorized'}), 403
        
    data = request.get_json() or {}
    class_id = data.get('class_id')
    mapel_id = data.get('mapel_id')
    kkm = float(data.get('kkm', 75.0))
    max_limit = float(data.get('max_limit', 98.0))
    students = data.get('students', [])
    
    if not class_id or not mapel_id or not students:
        return jsonify({'success': False, 'message': 'Data tidak lengkap / belum ada siswa'}), 400
        
    processed = []
    for s in students:
        nh1 = float(s.get('nh1') or 0.0)
        nh2 = float(s.get('nh2') or 0.0)
        nh3 = float(s.get('nh3') or 0.0)
        raw_avg = round((nh1 + nh2 + nh3) / 3.0, 2)
        processed.append({
            'nisn': s.get('nisn'),
            'nh1': nh1,
            'nh2': nh2,
            'nh3': nh3,
            'raw_average': raw_avg
        })
        
    # AI Scaling Engine
    raw_values = [s['raw_average'] for s in processed]
    min_raw = min(raw_values) if raw_values else 0.0
    max_raw = max(raw_values) if raw_values else 100.0
    current_range = max_raw - min_raw if max_raw - min_raw > 0 else 1.0
    target_range = max_limit - kkm
    scale_factor = target_range / current_range
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    for s in processed:
        cursor.execute("SELECT id FROM students WHERE nisn=?", (s['nisn'],))
        st_row = cursor.fetchone()
        if st_row:
            student_id = st_row[0]
            new_avg = kkm + ((s['raw_average'] - min_raw) * scale_factor)
            ai_adjusted = round(min(new_avg, max_limit), 2)
            
            cursor.execute('''INSERT OR REPLACE INTO nilai_akhir (student_id, mapel_id, nilai_asli, nilai_ai, nh1, nh2, nh3) 
                              VALUES (?, ?, ?, ?, ?, ?, ?)''', 
                           (student_id, mapel_id, s['raw_average'], ai_adjusted, s['nh1'], s['nh2'], s['nh3']))
        
    conn.commit()
    conn.close()
    
    return jsonify({'success': True, 'message': 'Nilai berhasil disimpan & diranking otomatis!'})

@app.route('/api/guru/upload', methods=['POST'])
def guru_upload_excel():
    if flask_session.get('role') != 'guru': return jsonify({'error': 'Unauthorized'}), 403
    if 'file' not in request.files: return jsonify({'error': 'File tidak ditemukan'}), 400
    
    file = request.files['file']
    class_id = request.form.get('class_id')
    mapel_id = request.form.get('mapel_id')
    kkm = float(request.form.get('kkm', 75.0))
    max_limit = float(request.form.get('maxLimit', 100.0))
    
    if not class_id or not mapel_id: return jsonify({'error': 'Kelas dan Mapel harus dipilih'}), 400
    
    try:
        import pandas as pd
        df = pd.read_excel(file)
        processed_data = []
        
        for index, row in df.iterrows():
            name = None
            for col in df.columns:
                col_str = str(col).lower()
                if any(x in col_str for x in ['nama', 'name', 'siswa', 'student']):
                    val = row[col]
                    if pd.notna(val) and not str(val).strip().isdigit():
                        name = str(val).strip()
                        break
            
            if not name and index < 50: 
                name = f'Siswa_{index + 1}'
            
            nisn = None
            for col in df.columns:
                col_str = str(col).lower()
                if 'nisn' in col_str or 'nis' in col_str:
                    val = row[col]
                    if pd.notna(val):
                        nisn = str(val).strip()
                        break

            gender = None
            for col in df.columns:
                col_str = str(col).lower()
                if any(x in col_str for x in ['l/p', 'gender', 'jenis kelamin', 'sex', 'lp']):
                    val = row[col]
                    if pd.notna(val):
                        val_str = str(val).strip().upper()
                        gender = 'P' if any(g in val_str for g in ['P', 'W', 'FEMALE', 'PR']) else 'L'
                        break
            
            if not gender:
                gender = 'L' if index % 2 == 0 else 'P'
            
            numeric_data = []
            for col in df.columns:
                col_lower = str(col).lower()
                if any(x in col_lower for x in ['nama', 'name', 'siswa', 'student', 'nisn', 'nis', 'id', 'no', 'l/p', 'gender', 'jenis kelamin', 'lp']):
                    continue
                try:
                    val = float(row[col]) if pd.notna(row[col]) else None
                    if val is not None: numeric_data.append(val)
                except: 
                    continue
                
            if len(numeric_data) < 1: continue
            
            nh1 = numeric_data[0] if len(numeric_data) > 0 else 0.0
            nh2 = numeric_data[1] if len(numeric_data) > 1 else nh1
            nh3 = numeric_data[2] if len(numeric_data) > 2 else nh1
            raw_avg = round((nh1 + nh2 + nh3) / 3.0, 2)
            
            processed_data.append({
                'id': index + 1, 
                'name': name, 
                'raw_average': raw_avg, 
                'nisn': nisn, 
                'gender': gender,
                'nh1': nh1,
                'nh2': nh2,
                'nh3': nh3
            })
            
        if not processed_data: return jsonify({'error': 'Tidak ada data nilai yang valid'}), 400
        
        raw_values = [s['raw_average'] for s in processed_data]
        min_raw = min(raw_values)
        max_raw = max(raw_values)
        current_range = max_raw - min_raw if max_raw - min_raw > 0 else 1.0
        target_range = max_limit - kkm
        scale_factor = target_range / current_range
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        for idx, student in enumerate(processed_data):
            nisn_val = student['nisn'] or f"NISN_{class_id}_{idx + 1}"
            new_avg = kkm + ((student['raw_average'] - min_raw) * scale_factor)
            ai_adjusted = round(min(new_avg, max_limit), 2)
            
            cursor.execute("SELECT id FROM students WHERE nisn=?", (nisn_val,))
            st_row = cursor.fetchone()
            if st_row:
                student_id = st_row[0]
                cursor.execute("UPDATE students SET nama_siswa=?, class_id=?, gender=? WHERE id=?", 
                               (student['name'], class_id, student['gender'], student_id))
            else:
                cursor.execute("INSERT INTO students (nama_siswa, nisn, class_id, gender) VALUES (?, ?, ?, ?)", 
                               (student['name'], nisn_val, class_id, student['gender']))
                student_id = cursor.lastrowid
                
            cursor.execute('''INSERT OR REPLACE INTO nilai_akhir (student_id, mapel_id, nilai_asli, nilai_ai, nh1, nh2, nh3) 
                              VALUES (?, ?, ?, ?, ?, ?, ?)''', 
                           (student_id, mapel_id, student['raw_average'], ai_adjusted, student['nh1'], student['nh2'], student['nh3']))
            
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': f'Berhasil memproses & meranking {len(processed_data)} siswa!'})
        
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# Perbaikan Query menggunakan LEFT JOIN agar siswa yang belum memiliki nilai tetap tampil di halaman Guru
@app.route('/api/guru/rekap', methods=['GET'])
def get_rekap():
    if flask_session.get('role') != 'guru': return jsonify({'error': 'Unauthorized'}), 403
    
    class_id = request.args.get('class_id')
    mapel_id = request.args.get('mapel_id')
    if not class_id or not mapel_id: return jsonify([])
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT s.nama_siswa, s.nisn, 
               COALESCE(n.nilai_asli, 0.0) as nilai_asli, 
               COALESCE(n.nilai_ai, 0.0) as nilai_ai,
               COALESCE(n.nilai_ai - n.nilai_asli, 0.0) as adj_point, 
               s.gender,
               COALESCE(n.nh1, 0.0) as nh1, 
               COALESCE(n.nh2, 0.0) as nh2, 
               COALESCE(n.nh3, 0.0) as nh3,
               RANK() OVER (ORDER BY COALESCE(n.nilai_ai, 0.0) DESC) as ranking
        FROM students s
        LEFT JOIN nilai_akhir n ON s.id = n.student_id AND n.mapel_id = ?
        WHERE s.class_id = ?
        ORDER BY s.nama_siswa ASC
    ''', (mapel_id, class_id))
    
    data = [{
        'nama_siswa': r[0], 
        'nisn': r[1], 
        'nilai_asli': r[2], 
        'nilai_ai': r[3], 
        'adj_point': round(r[4], 2), 
        'gender': r[5],
        'nh1': r[6],
        'nh2': r[7],
        'nh3': r[8],
        'ranking': r[9]
    } for r in cursor.fetchall()]
    
    conn.close()
    return jsonify(data)

@app.route('/api/guru/rekap_kelas', methods=['GET'])
def get_rekap_kelas():
    if flask_session.get('role') != 'guru': return jsonify({'error': 'Unauthorized'}), 403
    class_id = request.args.get('class_id')
    if not class_id: return jsonify({'error': 'class_id wajib diisi'}), 400
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("SELECT id, nama_mapel, COALESCE(kode_mapel, nama_mapel) as kode_mapel FROM mapels ORDER BY id ASC")
    mapels = [{'id': r[0], 'nama_mapel': r[1], 'kode_mapel': r[2] if r[2] else r[1]} for r in cursor.fetchall()]
    
    cursor.execute("SELECT id, nama_siswa, nisn, gender FROM students WHERE class_id = ?", (class_id,))
    students_raw = cursor.fetchall()
    
    if not students_raw:
        conn.close()
        return jsonify({'mapels': mapels, 'students': []})
        
    cursor.execute('''
        SELECT student_id, mapel_id, nilai_ai FROM nilai_akhir 
        WHERE student_id IN (SELECT id FROM students WHERE class_id = ?)
    ''', (class_id,))
    grades_raw = cursor.fetchall()
    conn.close()
    
    grades_map = {}
    for s_id, m_id, val in grades_raw:
        grades_map[(s_id, m_id)] = val
        
    student_records = []
    for s_id, nama, nisn, gender in students_raw:
        scores = {}
        total_sum = 0.0
        score_count = 0
        
        for m in mapels:
            val = grades_map.get((s_id, m['id']), None)
            scores[str(m['id'])] = val
            if val is not None:
                total_sum += val
                score_count += 1
                
        avg = round(total_sum / score_count, 2) if score_count > 0 else 0.0
        
        student_records.append({
            'student_id': s_id,
            'nama_siswa': nama,
            'nisn': nisn,
            'gender': gender,
            'scores': scores,
            'total_sum': round(total_sum, 2),
            'average': avg
        })
        
    student_records.sort(key=lambda x: x['total_sum'], reverse=True)
    for idx, rec in enumerate(student_records):
        rec['ranking'] = idx + 1
        
    student_records.sort(key=lambda x: x['nama_siswa'])
    
    return jsonify({
        'mapels': mapels,
        'students': student_records
    })


# ==========================================
# 📄 DOWNLOAD & REPORT EXPORT API (EXCEL & PDF)
# ==========================================
@app.route('/api/guru/export/excel', methods=['GET'])
def export_excel():
    if flask_session.get('role') != 'guru':
        return jsonify({'error': 'Unauthorized'}), 403
        
    class_id = request.args.get('class_id')
    mapel_id = request.args.get('mapel_id')
    tipe_guru = flask_session.get('tipe_guru')
    
    if not class_id:
        return "Parameter tidak lengkap", 400
        
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT nama_kelas FROM classes WHERE id=?", (class_id,))
    class_row = cursor.fetchone()
    nama_kelas = class_row[0] if class_row else f"Kelas_{class_id}"
    
    nama_guru = flask_session.get('nama_guru', 'Ronaldo, S.Pd., Gr.')
    nip_guru = flask_session.get('nip', '1234567890')
    
    if tipe_guru == 'kelas' and (not mapel_id or mapel_id == 'semua'):
        cursor.execute("SELECT id, nama_mapel, COALESCE(kode_mapel, nama_mapel) as kode_mapel FROM mapels ORDER BY id ASC")
        mapels = [{'id': r[0], 'nama_mapel': r[1], 'kode_mapel': r[2] if r[2] else r[1]} for r in cursor.fetchall()]
        
        cursor.execute("SELECT id, nama_siswa, nisn, gender FROM students WHERE class_id = ?", (class_id,))
        students_raw = cursor.fetchall()
        
        cursor.execute('''
            SELECT student_id, mapel_id, nilai_ai FROM nilai_akhir 
            WHERE student_id IN (SELECT id FROM students WHERE class_id = ?)
        ''', (class_id,))
        grades_raw = cursor.fetchall()
        conn.close()
        
        grades_map = {}
        for s_id, m_id, val in grades_raw:
            grades_map[(s_id, m_id)] = val
            
        student_records = []
        for s_id, nama, nisn, gender in students_raw:
            scores = {}
            total_sum = 0.0
            score_count = 0
            for m in mapels:
                val = grades_map.get((s_id, m['id']), 0)
                scores[m['id']] = val
                total_sum += val
                if val > 0: score_count += 1
            avg = round(total_sum / score_count, 2) if score_count > 0 else 0.0
            student_records.append({
                'nama': nama, 'nisn': nisn, 'gender': gender, 'scores': scores, 'total_sum': total_sum, 'avg': avg
            })
            
        student_records.sort(key=lambda x: x['total_sum'], reverse=True)
        for idx, rec in enumerate(student_records): rec['rank'] = idx + 1
        student_records.sort(key=lambda x: x['nama'])
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Rekap Nilai Siswa"
            ws.views.sheetView[0].showGridLines = True

            font_title = Font(name='Arial', size=16, bold=True, color='2D5F9E')
            font_sub = Font(name='Arial', size=12, bold=True, color='2D5F9E')
            font_meta = Font(name='Arial', size=10, bold=True)
            font_header = Font(name='Arial', size=10, bold=True, color='FFFFFF')
            font_data = Font(name='Arial', size=10)
            font_bold = Font(name='Arial', size=10, bold=True)
            
            fill_header = PatternFill(start_color='A0AEC0', end_color='A0AEC0', fill_type='solid')
            fill_footer = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            
            border_thin = Side(border_style="thin", color="CBD5E1")
            cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

            ws.merge_cells('A1:N1')
            ws['A1'] = "UPT SDN 005 PETAPAHAN JAYA"
            ws['A1'].font = font_title
            ws['A1'].alignment = Alignment(horizontal='center')
            
            ws.merge_cells('A2:N2')
            ws['A2'] = f"REKAPITULASI NILAI SISWA KELAS {nama_kelas.upper()}"
            ws['A2'].font = font_sub
            ws['A2'].alignment = Alignment(horizontal='center')
            
            ws['A4'] = "TAHUN"; ws['B4'] = ": 2025/2026 GENAP"
            ws['A5'] = "GURU KELAS"; ws['B5'] = f": {nama_guru}"
            ws['A6'] = "NIP"; ws['B6'] = f": {nip_guru}"
            
            ws['G4'] = "SISWA"; ws['H4'] = f": {len(student_records)}"
            ws['G5'] = "LK"; ws['H5'] = f": {sum(1 for s in student_records if s['gender'] == 'L'):02d}"
            ws['G6'] = "PR"; ws['H6'] = f": {sum(1 for s in student_records if s['gender'] == 'P'):02d}"
            
            headers = ["No.", "NISN", "Nama", "L/P"] + [m['kode_mapel'].upper() for m in mapels] + ["JUMLAH", "RATA-RATA", "RANK"]
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=9, column=col_idx)
                cell.value = h
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = cell_border

            row_start = 10
            mapel_sums = {m['id']: 0.0 for m in mapels}
            total_sum_all = 0.0
            
            for idx, r in enumerate(student_records):
                cur_row = row_start + idx
                total_sum_all += r['total_sum']
                
                data_row = [idx + 1, r['nisn'], r['nama'], r['gender']]
                for m in mapels:
                    score = r['scores'].get(m['id'], 0)
                    data_row.append(score if score > 0 else "")
                    mapel_sums[m['id']] += score
                data_row += [r['total_sum'], r['avg'], r['rank']]
                
                for col_idx, val in enumerate(data_row, 1):
                    cell = ws.cell(row=cur_row, column=col_idx)
                    cell.value = val
                    cell.font = font_data
                    cell.border = cell_border
                    if col_idx in [1, 2, 4, len(headers)]:
                        cell.alignment = Alignment(horizontal='center')
                    elif col_idx == 3:
                        cell.alignment = Alignment(horizontal='left')
                    else:
                        cell.alignment = Alignment(horizontal='right')

            # Footer
            avg_row = row_start + len(student_records)
            ws.merge_cells(start_row=avg_row, start_column=1, end_row=avg_row, end_column=4)
            avg_label = ws.cell(row=avg_row, column=1)
            avg_label.value = "RATA-RATA"
            avg_label.font = font_bold
            avg_label.alignment = Alignment(horizontal='center')
            avg_label.fill = fill_footer
            
            for c in range(1, 5):
                ws.cell(row=avg_row, column=c).border = cell_border
                ws.cell(row=avg_row, column=c).fill = fill_footer

            col_c = 5
            for m in mapels:
                cell = ws.cell(row=avg_row, column=col_c)
                cell.value = round(mapel_sums[m['id']] / len(student_records), 2) if student_records else 0.0
                cell.font = font_bold
                cell.border = cell_border
                cell.fill = fill_footer
                col_c += 1
                
            cell_total_sum = ws.cell(row=avg_row, column=col_c)
            cell_total_sum.value = round(total_sum_all / len(student_records), 2) if student_records else 0.0
            cell_total_sum.font = font_bold
            cell_total_sum.border = cell_border
            cell_total_sum.fill = fill_footer
            
            ws.cell(row=avg_row, column=col_c+1).border = cell_border
            ws.cell(row=avg_row, column=col_c+1).fill = fill_footer
            ws.cell(row=avg_row, column=col_c+2).border = cell_border
            ws.cell(row=avg_row, column=col_c+2).fill = fill_footer

            sig_start = avg_row + 3
            sig_col = len(headers) - 2
            ws.cell(row=sig_start, column=sig_col).value = f"Petapahan Jaya, {format_indo_date()}"
            ws.cell(row=sig_start, column=sig_col).font = font_meta
            ws.cell(row=sig_start+1, column=sig_col).value = "Guru Kelas,"
            ws.cell(row=sig_start+1, column=sig_col).font = font_meta
            ws.cell(row=sig_start+5, column=sig_col).value = nama_guru
            ws.cell(row=sig_start+5, column=sig_col).font = Font(name='Arial', size=10, bold=True, underline="single")
            ws.cell(row=sig_start+6, column=sig_col).value = f"NIP. {nip_guru}"
            ws.cell(row=sig_start+6, column=sig_col).font = font_meta

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=f"Rekap_Siswa_{nama_kelas}.xlsx")
        except Exception as e:
            print(f"Guru Kelas Excel failed: {e}")
            return "Excel Generation Error", 500

    else:
        # Jika mapel_id didefinisikan, cetak rekap per mapel tunggal
        cursor.execute('''
            SELECT s.nama_siswa, s.nisn, n.nilai_asli, n.nilai_ai,
                   (n.nilai_ai - n.nilai_asli) as adj_point, s.gender,
                   RANK() OVER (ORDER BY n.nilai_ai DESC) as ranking,
                   n.nh1, n.nh2, n.nh3
            FROM students s
            JOIN nilai_akhir n ON s.id = n.student_id
            WHERE s.class_id = ? AND n.mapel_id = ?
            ORDER BY ranking ASC
        ''', (class_id, mapel_id))
        rows = cursor.fetchall()
        
        cursor.execute("SELECT nama_mapel FROM mapels WHERE id=?", (mapel_id,))
        mapel_row = cursor.fetchone()
        nama_mapel = mapel_row[0] if mapel_row else f"Mapel_{mapel_id}"
        conn.close()
        
        siswa_count = len(rows)
        l_count = sum(1 for r in rows if r[5] == 'L')
        p_count = sum(1 for r in rows if r[5] == 'P')
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Rekapitulasi Nilai"
            ws.views.sheetView[0].showGridLines = True

            font_title = Font(name='Arial', size=16, bold=True, color='2D5F9E')
            font_sub = Font(name='Arial', size=12, bold=True, color='2D5F9E')
            font_meta = Font(name='Arial', size=10, bold=True)
            font_header = Font(name='Arial', size=10, bold=True, color='FFFFFF')
            font_data = Font(name='Arial', size=10)
            font_bold = Font(name='Arial', size=10, bold=True)
            font_under = Font(name='Arial', size=10, bold=True, underline="single")
            
            fill_header = PatternFill(start_color='A0AEC0', end_color='A0AEC0', fill_type='solid')
            fill_footer = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            
            align_center = Alignment(horizontal='center', vertical='center')
            align_left = Alignment(horizontal='left', vertical='center')
            align_right = Alignment(horizontal='right', vertical='center')
            
            border_thin = Side(border_style="thin", color="CBD5E1")
            cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

            ws.merge_cells('A1:K1')
            ws['A1'] = "UPT SDN 005 PETAPAHAN JAYA"
            ws['A1'].font = font_title
            ws['A1'].alignment = align_center
            
            ws.merge_cells('A2:K2')
            ws['A2'] = f"REKAPITULASI NILAI {nama_mapel.upper()} PER KELAS"
            ws['A2'].font = font_sub
            ws['A2'].alignment = align_center
            
            ws['A4'] = "KELAS"; ws['B4'] = f": {nama_kelas}"
            ws['A5'] = "TAHUN"; ws['B5'] = f": 2025/2026 GENAP"
            ws['A6'] = "GURU MAPEL"; ws['B6'] = f": {nama_guru}"
            ws['A7'] = "NIP"; ws['B7'] = f": {nip_guru}"
            
            ws['G4'] = "SISWA"; ws['H4'] = f": {siswa_count}"
            ws['G5'] = "L"; ws['H5'] = f": {l_count:02d}"
            ws['G6'] = "P"; ws['H6'] = f": {p_count:02d}"
            
            for r in range(4, 8):
                ws[f'A{r}'].font = font_meta
                ws[f'B{r}'].font = font_meta
                ws[f'G{r}'].font = font_meta
                ws[f'H{r}'].font = font_meta

            headers = ["No.", "NISN", "Nama", "L/P", "NH-1", "NH-2", "NH-3", "Nilai Asli", "Nilai", "Adj. Point", "Rank"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=9, column=col_idx)
                cell.value = header
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = cell_border

            sum_nh1, sum_nh2, sum_nh3, sum_asli, sum_ai, sum_adj = 0, 0, 0, 0, 0, 0
            row_start = 10
            
            for idx, r in enumerate(rows):
                cur_row = row_start + idx
                val_nh1 = r[7] if r[7] is not None else r[2]
                val_nh2 = r[8] if r[8] is not None else r[2]
                val_nh3 = r[9] if r[9] is not None else r[2]
                
                sum_nh1 += val_nh1
                sum_nh2 += val_nh2
                sum_nh3 += val_nh3
                sum_asli += r[2]
                sum_ai += r[3]
                sum_adj += r[4]
                
                data_row = [idx + 1, r[1], r[0], r[5], val_nh1, val_nh2, val_nh3, r[2], int(r[3]), int(r[4]), r[6]]
                for col_idx, val in enumerate(data_row, 1):
                    cell = ws.cell(row=cur_row, column=col_idx)
                    cell.value = val
                    cell.font = font_data
                    cell.border = cell_border
                    if col_idx in [1, 2, 4, 11]:
                        cell.alignment = align_center
                    elif col_idx == 3:
                        cell.alignment = align_left
                    else:
                        cell.alignment = align_right

            avg_row = row_start + len(rows)
            ws.merge_cells(start_row=avg_row, start_column=1, end_row=avg_row, end_column=4)
            avg_label_cell = ws.cell(row=avg_row, column=1)
            avg_label_cell.value = "RATA-RATA"
            avg_label_cell.font = font_bold
            avg_label_cell.alignment = align_center
            avg_label_cell.fill = fill_footer
            avg_label_cell.border = cell_border
            
            for c in range(1, 5):
                ws.cell(row=avg_row, column=c).border = cell_border
                ws.cell(row=avg_row, column=c).fill = fill_footer

            avg_values = {
                5: f"{sum_nh1/siswa_count:.2f}", 6: f"{sum_nh2/siswa_count:.2f}", 7: f"{sum_nh3/siswa_count:.2f}",
                8: f"{sum_asli/siswa_count:.2f}", 9: f"{sum_ai/siswa_count:.2f}", 10: f"{sum_adj/siswa_count:.2f}"
            }
            
            for col_idx, avg_val in avg_values.items():
                cell = ws.cell(row=avg_row, column=col_idx)
                cell.value = float(avg_val)
                cell.font = font_bold
                cell.alignment = align_right
                cell.border = cell_border
                cell.fill = fill_footer
                
            ws.cell(row=avg_row, column=11).border = cell_border
            ws.cell(row=avg_row, column=11).fill = fill_footer

            sig_start = avg_row + 3
            ws.cell(row=sig_start, column=8).value = f"Petapahan Jaya, {format_indo_date()}"
            ws.cell(row=sig_start, column=8).font = font_meta
            ws.cell(row=sig_start+1, column=8).value = "Guru Mata Pelajaran,"
            ws.cell(row=sig_start+1, column=8).font = font_meta
            ws.cell(row=sig_start+5, column=8).value = nama_guru
            ws.cell(row=sig_start+5, column=8).font = font_under
            ws.cell(row=sig_start+6, column=8).value = f"NIP. {nip_guru}"
            ws.cell(row=sig_start+6, column=8).font = font_meta

            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.row in [1, 2]: continue 
                    if cell.value: max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
                
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=f"Rekap_Nilai_{nama_kelas}_{nama_mapel}.xlsx")
        except Exception as e:
            print(f"Guru Mapel Excel failed: {e}")
            return "Excel Generation Error", 500

@app.route('/api/guru/export/pdf', methods=['GET'])
def export_pdf():
    if flask_session.get('role') != 'guru':
        return jsonify({'error': 'Unauthorized'}), 403
        
    class_id = request.args.get('class_id')
    mapel_id = request.args.get('mapel_id')
    tipe_guru = flask_session.get('tipe_guru')
    
    if not class_id:
        return "Parameter tidak lengkap", 400
        
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT nama_kelas FROM classes WHERE id=?", (class_id,))
    class_row = cursor.fetchone()
    nama_kelas = class_row[0] if class_row else f"Kelas_{class_id}"
    
    nama_guru = flask_session.get('nama_guru', 'Ronaldo, S.Pd., Gr.')
    nip_guru = flask_session.get('nip', '1234567890')
    
    # ----------------- PRINT LAYOUT GURU KELAS -----------------
    if tipe_guru == 'kelas' and (not mapel_id or mapel_id == 'semua'):
        cursor.execute("SELECT id, nama_mapel, COALESCE(kode_mapel, nama_mapel) as kode_mapel FROM mapels ORDER BY id ASC")
        mapels = [{'id': r[0], 'nama_mapel': r[1], 'kode_mapel': r[2] if r[2] else r[1]} for r in cursor.fetchall()]
        
        cursor.execute("SELECT id, nama_siswa, nisn, gender FROM students WHERE class_id = ?", (class_id,))
        students_raw = cursor.fetchall()
        
        cursor.execute('''
            SELECT student_id, mapel_id, nilai_ai FROM nilai_akhir 
            WHERE student_id IN (SELECT id FROM students WHERE class_id = ?)
        ''', (class_id,))
        grades_raw = cursor.fetchall()
        conn.close()
        
        grades_map = {}
        for s_id, m_id, val in grades_raw:
            grades_map[(s_id, m_id)] = val
            
        student_records = []
        for s_id, nama, nisn, gender in students_raw:
            scores = {}
            total_sum = 0.0
            score_count = 0
            for m in mapels:
                val = grades_map.get((s_id, m['id']), 0)
                scores[m['id']] = val
                total_sum += val
                if val > 0: score_count += 1
            avg = round(total_sum / score_count, 2) if score_count > 0 else 0.0
            student_records.append({
                'nama': nama, 'nisn': nisn, 'gender': gender, 'scores': scores, 'total_sum': total_sum, 'avg': avg
            })
            
        student_records.sort(key=lambda x: x['total_sum'], reverse=True)
        for idx, rec in enumerate(student_records): rec['rank'] = idx + 1
        student_records.sort(key=lambda x: x['nama'])
        
        siswa_count = len(student_records)
        l_count = sum(1 for s in student_records if s['gender'] == 'L')
        p_count = sum(1 for s in student_records if s['gender'] == 'P')
        
        mapel_averages = []
        for m in mapels:
            total_m = sum(s['scores'].get(m['id'], 0) for s in student_records)
            mapel_averages.append(round(total_m / siswa_count, 2) if siswa_count else 0.0)
            
        avg_total_all = round(sum(s['total_sum'] for s in student_records) / siswa_count, 2) if siswa_count else 0.0

        html_print_layout = """
        <!DOCTYPE html>
        <html>
        <head>
            <title>Rekapitulasi Nilai Kelas - UPT SDN 005 Petapahan Jaya</title>
            <style>
                body { font-family: 'Arial', sans-serif; padding: 20px; color: #1e293b; background: #fff; }
                .header-print { text-align: center; margin-bottom: 25px; }
                .header-print h1 { margin: 0; font-size: 22px; font-weight: bold; color: #1e3a8a; text-transform: uppercase; }
                .header-print h2 { margin: 5px 0 0 0; font-size: 14px; font-weight: bold; color: #1e3a8a; }
                
                .meta-section { display: flex; justify-content: space-between; margin-bottom: 20px; font-size: 12px; font-weight: bold; line-height: 1.8; }
                .meta-col { width: 45%; }
                .meta-row { display: flex; }
                .meta-label { width: 120px; text-transform: uppercase; }
                .meta-val { flex: 1; }
                
                .print-table { width: 100%; border-collapse: collapse; margin-top: 10px; font-size: 10px; }
                .print-table th, .print-table td { border: 1px solid #94a3b8; padding: 6px 8px; }
                .print-table th { background-color: #cbd5e1; font-weight: bold; text-align: center; color: #0f172a; text-transform: uppercase; }
                
                .text-center { text-align: center; }
                .text-left { text-align: left; }
                .text-right { text-align: right; }
                
                .row-stripe:nth-child(even) { background-color: #f8fafc; }
                .footer-avg { background-color: #f1f5f9; font-weight: bold; }
                
                .no-print-area { margin-bottom: 25px; background: #fffbeb; border: 1px solid #f59e0b; padding: 15px; border-radius: 8px; font-size: 13px; }
                .no-print-area button { background: #1e3a8a; color: white; border: none; padding: 10px 20px; border-radius: 6px; cursor: pointer; font-weight: bold; }
                
                .signature-container { margin-top: 40px; display: flex; justify-content: flex-end; }
                .signature-box { width: 300px; text-align: left; font-size: 13px; font-weight: bold; }
                
                @media print {
                    .no-print-area { display: none; }
                    body { padding: 0; }
                }
            </style>
        </head>
        <body>
            <div class="no-print-area">
                <button onclick="window.print()">Cetak / Simpan PDF</button>
            </div>

            <div class="header-print">
                <h1>UPT SDN 005 PETAPAHAN JAYA</h1>
                <h2>REKAPITULASI NILAI SISWA KELAS {{ kelas }}</h2>
            </div>

            <div class="meta-section">
                <div class="meta-col">
                    <div class="meta-row"><div class="meta-label">Tahun</div><div class="meta-val">: 2025/2026 GENAP</div></div>
                    <div class="meta-row"><div class="meta-label">Guru Kelas</div><div class="meta-val">: {{ guru_kelas }}</div></div>
                    <div class="meta-row"><div class="meta-label">NIP</div><div class="meta-val">: {{ nip }}</div></div>
                </div>
                <div class="meta-col" style="padding-left: 100px;">
                    <div class="meta-row"><div class="meta-label">Siswa</div><div class="meta-val">: {{ siswa_count }}</div></div>
                    <div class="meta-row"><div class="meta-label">LK</div><div class="meta-val">: {{ format_count(l_count) }}</div></div>
                    <div class="meta-row"><div class="meta-label">PR</div><div class="meta-val">: {{ format_count(p_count) }}</div></div>
                </div>
            </div>

            <table class="print-table">
                <thead>
                    <tr>
                        <th width="4%">No.</th>
                        <th>NISN</th>
                        <th>Nama</th>
                        <th width="5%">L/P</th>
                        {% for m in mapels %}
                        <th>{{ m.kode_mapel | upper }}</th>
                        {% endfor %}
                        <th>JUMLAH</th>
                        <th>RATA-RATA</th>
                        <th width="5%">RANK</th>
                    </tr>
                </thead>
                <tbody>
                    {% for r in student_records %}
                    <tr class="row-stripe">
                        <td class="text-center">{{ loop.index }}</td>
                        <td class="text-center">{{ r.nisn }}</td>
                        <td class="text-left" style="font-weight: 500;">{{ r.nama }}</td>
                        <td class="text-center">{{ r.gender }}</td>
                        {% for m in mapels %}
                        <td class="text-right">
                            {% if r.scores[m.id] > 0 %}
                                {{ r.scores[m.id] | int }}
                            {% else %}
                                -
                            {% endif %}
                        </td>
                        {% endfor %}
                        <td class="text-right" style="font-weight: bold;">{{ r.total_sum | int }}</td>
                        <td class="text-right">{{ "%.2f" | format(r.avg) }}</td>
                        <td class="text-center" style="font-weight: bold;">{{ r.rank }}</td>
                    </tr>
                    {% endfor %}
                    <tr class="footer-avg">
                        <td colspan="4" class="text-center">RATA-RATA</td>
                        {% for avg in mapel_averages %}
                        <td class="text-right">{{ "%.2f" | format(avg) }}</td>
                        {% endfor %}
                        <td class="text-right">{{ "%.2f" | format(avg_total_all) }}</td>
                        <td></td>
                        <td></td>
                    </tr>
                </tbody>
            </table>

            <div class="signature-container">
                <div class="signature-box">
                    <p style="margin-bottom: 2px;">Petapahan Jaya, {{ date_indo }}</p>
                    <p style="margin-top: 0; margin-bottom: 60px;">Guru Kelas,</p>
                    <p style="margin-bottom: 2px; font-weight: bold; text-decoration: underline;">{{ guru_kelas }}</p>
                    <p style="margin-top: 0;">NIP. {{ nip }}</p>
                </div>
            </div>

            <script>
                window.onload = function() {
                    setTimeout(function() {
                        window.print();
                    }, 500);
                };
            </script>
        </body>
        </html>
        """
        
        def format_count(value):
            try: return f"{int(value):02d}"
            except: return value

        return render_template_string(
            html_print_layout,
            student_records=student_records,
            mapels=mapels,
            mapel_averages=mapel_averages,
            avg_total_all=avg_total_all,
            kelas=nama_kelas,
            guru_kelas=nama_guru,
            nip=nip_guru,
            siswa_count=siswa_count,
            l_count=l_count,
            p_count=p_count,
            format_count=format_count,
            date_indo=format_indo_date()
        )

    # ----------------- PRINT LAYOUT GURU MAPEL / SINGLE SUBJECT -----------------
    else:
        cursor.execute('''
            SELECT s.nama_siswa, s.nisn, n.nilai_asli, n.nilai_ai,
                   (n.nilai_ai - n.nilai_asli) as adj_point, s.gender,
                   RANK() OVER (ORDER BY n.nilai_ai DESC) as ranking,
                   n.nh1, n.nh2, n.nh3
            FROM students s
            JOIN nilai_akhir n ON s.id = n.student_id
            WHERE s.class_id = ? AND n.mapel_id = ?
            ORDER BY ranking ASC
        ''', (class_id, mapel_id))
        rows = cursor.fetchall()
        
        cursor.execute("SELECT nama_mapel FROM mapels WHERE id=?", (mapel_id,))
        mapel_row = cursor.fetchone()
        nama_mapel = mapel_row[0] if mapel_row else f"Mapel_{mapel_id}"
        conn.close()
        
        siswa_count = len(rows)
        l_count = sum(1 for r in rows if r[5] == 'L')
        p_count = sum(1 for r in rows if r[5] == 'P')
        
        avg_asli = round(sum(r[2] for r in rows) / siswa_count, 2) if siswa_count else 0
        avg_ai = round(sum(r[3] for r in rows) / siswa_count, 2) if siswa_count else 0
        avg_adj = round(sum(r[4] for r in rows) / siswa_count, 2) if siswa_count else 0
        avg_nh1 = round(sum((r[7] if r[7] is not None else r[2]) for r in rows) / siswa_count, 2) if siswa_count else 0
        avg_nh2 = round(sum((r[8] if r[8] is not None else r[2]) for r in rows) / siswa_count, 2) if siswa_count else 0
        avg_nh3 = round(sum((r[9] if r[9] is not None else r[2]) for r in rows) / siswa_count, 2) if siswa_count else 0

        html_print_layout = """
        <!DOCTYPE html>
        <html>
        <head>
            <title>Rekapitulasi Nilai Resmi - UPT SDN 005 Petapahan Jaya</title>
            <style>
                body { font-family: 'Arial', sans-serif; padding: 20px; color: #1e293b; background: #fff; }
                .header-print { text-align: center; margin-bottom: 25px; }
                .header-print h1 { margin: 0; font-size: 22px; font-weight: bold; color: #1e3a8a; text-transform: uppercase; }
                .header-print h2 { margin: 5px 0 0 0; font-size: 14px; font-weight: bold; color: #1e3a8a; letter-spacing: 1px; }
                
                .meta-section { display: flex; justify-content: space-between; margin-bottom: 20px; font-size: 12px; font-weight: bold; line-height: 1.8; }
                .meta-col { width: 45%; }
                .meta-row { display: flex; }
                .meta-label { width: 120px; text-transform: uppercase; }
                .meta-val { flex: 1; }
                
                .print-table { width: 100%; border-collapse: collapse; margin-top: 10px; font-size: 11px; }
                .print-table th, .print-table td { border: 1px solid #94a3b8; padding: 8px 10px; }
                .print-table th { background-color: #cbd5e1; font-weight: bold; text-align: center; color: #0f172a; text-transform: uppercase; }
                
                .text-center { text-align: center; }
                .text-left { text-align: left; }
                .text-right { text-align: right; }
                
                .row-stripe:nth-child(even) { background-color: #f8fafc; }
                .footer-avg { background-color: #f1f5f9; font-weight: bold; }
                
                .no-print-area { margin-bottom: 25px; background: #fffbeb; border: 1px solid #f59e0b; padding: 15px; border-radius: 8px; font-size: 13px; }
                .no-print-area button { background: #1e3a8a; color: white; border: none; padding: 10px 20px; border-radius: 6px; cursor: pointer; font-weight: bold; }
                
                .nh-col { display: table-cell; }
                
                .signature-container { margin-top: 40px; display: flex; justify-content: flex-end; }
                .signature-box { width: 300px; text-align: left; font-size: 13px; font-weight: bold; }
                
                @media print {
                    .no-print-area { display: none; }
                    body { padding: 0; }
                }
            </style>
        </head>
        <body>
            <div class="no-print-area">
                <label style="display:inline-flex; align-items:center; gap:8px; margin-right:20px; font-weight:bold; cursor:pointer;">
                    <input type="checkbox" id="toggleNHPrint" checked onchange="toggleNHColumnsPrint()"> Tampilkan Kolom NH & Nilai Asli
                </label>
                <button onclick="window.print()">Cetak / Simpan PDF</button>
            </div>

            <div class="header-print">
                <h1>UPT SDN 005 PETAPAHAN JAYA</h1>
                <h2>REKAPITULASI NILAI {{ mapel | upper }} PER KELAS</h2>
            </div>

            <div class="meta-section">
                <div class="meta-col">
                    <div class="meta-row"><div class="meta-label">Kelas</div><div class="meta-val">: {{ kelas }}</div></div>
                    <div class="meta-row"><div class="meta-label">Tahun</div><div class="meta-val">: 2025/2026 GENAP</div></div>
                    <div class="meta-row"><div class="meta-label">Guru Mapel</div><div class="meta-val">: {{ guru_mapel }}</div></div>
                    <div class="meta-row"><div class="meta-label">NIP</div><div class="meta-val">: {{ nip }}</div></div>
                </div>
                <div class="meta-col" style="padding-left: 100px;">
                    <div class="meta-row"><div class="meta-label">Siswa</div><div class="meta-val">: {{ siswa_count }}</div></div>
                    <div class="meta-row"><div class="meta-label">L</div><div class="meta-val">: {{ format_count(l_count) }}</div></div>
                    <div class="meta-row"><div class="meta-label">P</div><div class="meta-val">: {{ format_count(p_count) }}</div></div>
                </div>
            </div>

            <table class="print-table">
                <thead>
                    <tr>
                        <th width="4%">No.</th>
                        <th width="10%">NISN</th>
                        <th>Nama</th>
                        <th width="6%">L/P</th>
                        <th width="8%" class="nh-col">NH-1</th>
                        <th width="8%" class="nh-col">NH-2</th>
                        <th width="8%" class="nh-col">NH-3</th>
                        <th width="8%" class="nh-col">Nilai Asli</th>
                        <th width="8%">Nilai</th>
                        <th width="8%">Adj. Point</th>
                        <th width="6%">Rank</th>
                    </tr>
                </thead>
                <tbody>
                    {% for r in rows %}
                    <tr class="row-stripe">
                        <td class="text-center">{{ loop.index }}</td>
                        <td class="text-center">{{ r[1] }}</td>
                        <td class="text-left" style="font-weight: 500;">{{ r[0] }}</td>
                        <td class="text-center">{{ r[5] }}</td>
                        
                        <td class="text-right nh-col">{{ "%.1f" | format(r[7] if r[7] is not None else r[2]) }}</td>
                        <td class="text-right nh-col">{{ "%.1f" | format(r[8] if r[8] is not None else r[2]) }}</td>
                        <td class="text-right nh-col">{{ "%.1f" | format(r[9] if r[9] is not None else r[2]) }}</td>
                        <td class="text-right nh-col">{{ "%.1f" | format(r[2]) }}</td>
                        
                        <td class="text-right" style="font-weight: bold; color: #1e3a8a;">{{ r[3] | int }}</td>
                        <td class="text-right">{{ r[4] | int }}</td>
                        <td class="text-center" style="font-weight: bold;">{{ r[6] }}</td>
                    </tr>
                    {% endfor %}
                    <tr class="footer-avg">
                        <td colspan="4" class="text-center">RATA-RATA</td>
                        <td class="text-right nh-col">{{ "%.2f" | format(avg_nh1) }}</td>
                        <td class="text-right nh-col">{{ "%.2f" | format(avg_nh2) }}</td>
                        <td class="text-right nh-col">{{ "%.2f" | format(avg_nh3) }}</td>
                        <td class="text-right nh-col">{{ "%.2f" | format(avg_asli) }}</td>
                        <td class="text-right">{{ "%.2f" | format(avg_ai) }}</td>
                        <td class="text-right">{{ "%.2f" | format(avg_adj) }}</td>
                        <td></td>
                    </tr>
                </tbody>
            </table>

            <div class="signature-container">
                <div class="signature-box">
                    <p style="margin-bottom: 2px;">Petapahan Jaya, {{ date_indo }}</p>
                    <p style="margin-top: 0; margin-bottom: 60px;">Guru Mata Pelajaran,</p>
                    <p style="margin-bottom: 2px; font-weight: bold; text-decoration: underline;">{{ guru_mapel }}</p>
                    <p style="margin-top: 0;">NIP. {{ nip }}</p>
                </div>
            </div>

            <script>
                function toggleNHColumnsPrint() {
                    const show = document.getElementById('toggleNHPrint').checked;
                    const cols = document.querySelectorAll('.nh-col');
                    cols.forEach(el => {
                        el.style.display = show ? 'table-cell' : 'none';
                    });
                }
                
                window.onload = function() {
                    setTimeout(function() {
                        window.print();
                    }, 500);
                };
            </script>
        </body>
        </html>
        """
        
        def format_count(value):
            try: return f"{int(value):02d}"
            except: return value

        return render_template_string(
            html_print_layout, 
            rows=rows, 
            kelas=nama_kelas, 
            mapel=nama_mapel, 
            guru_mapel=nama_guru,
            nip=nip_guru,
            siswa_count=siswa_count,
            l_count=l_count,
            p_count=p_count,
            avg_asli=avg_asli,
            avg_ai=avg_ai,
            avg_adj=avg_adj,
            avg_nh1=avg_nh1,
            avg_nh2=avg_nh2,
            avg_nh3=avg_nh3,
            format_count=format_count,
            date_indo=format_indo_date()
        )


if __name__ == '__main__':
    app.run(debug=True, host='127.0.0.1', port=5000)